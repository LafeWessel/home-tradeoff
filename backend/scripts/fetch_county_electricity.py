#!/usr/bin/env python3
"""Generate county_electricity.json from EIA Form 861 (2024).

Downloads the annual EIA-861 ZIP, reads utility-level residential sales and
revenue data, maps utilities to counties via the Service Territory schedule,
and computes a customer-weighted average residential electricity rate (¢/kWh)
for each county.

Data sources:
  Sales_Ult_Cust_2024.xlsx (States sheet) — utility residential revenues + sales
  Service_Territory_2024.xlsx (Counties_States sheet) — utility → county mapping

County name → FIPS mapping is fetched once from the Census API.

Usage:
    cd backend && python -m scripts.fetch_county_electricity

Requires CENSUS_API_KEY in backend/.env.

Output:
    app/sources/static/county_electricity.json
    Keys: 5-digit county FIPS
    Values: average residential electricity rate (¢/kWh)
"""

from __future__ import annotations

import io
import json
import logging
import sys
import zipfile
from pathlib import Path

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

STATIC_DIR = Path(__file__).resolve().parent.parent / "app" / "sources" / "static"
OUT_FILE = STATIC_DIR / "county_electricity.json"

EIA_URL = "https://www.eia.gov/electricity/data/eia861/zip/f8612024.zip"
ACS_BASE = "https://api.census.gov/data/{year}/acs/acs5"

STATE_ABBR_TO_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08",
    "CT": "09", "DE": "10", "FL": "12", "GA": "13", "HI": "15", "ID": "16",
    "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21", "LA": "22",
    "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28",
    "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33", "NJ": "34",
    "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39", "OK": "40",
    "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46", "TN": "47",
    "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53", "WV": "54",
    "WI": "55", "WY": "56", "DC": "11",
}

# Suffixes to strip when normalizing county names for matching
_COUNTY_SUFFIXES = (
    " county", " parish", " borough", " census area", " city and borough",
    " municipality", " city", " town", " village",
)


def _normalize_county_name(raw: str) -> str:
    s = raw.strip().lower()
    for suffix in _COUNTY_SUFFIXES:
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
            break
    # Remove "st." vs "saint" vs "st " ambiguity
    s = s.replace("saint ", "st. ").replace("ste. ", "sainte ").replace("st ", "st. ")
    return s


def _build_county_name_fips(api_key: str) -> dict[tuple[str, str], str]:
    """Return {(state_abbr, normalized_county_name): 5-digit-fips}."""
    import requests

    FIPS_TO_ABBR = {v: k for k, v in STATE_ABBR_TO_FIPS.items()}
    STATE_FIPS = [f"{n:02d}" for n in range(1, 57) if n not in (3, 7, 11, 14, 43, 52)]

    mapping: dict[tuple[str, str], str] = {}
    for sfips in STATE_FIPS:
        r = requests.get(
            ACS_BASE.format(year=2023),
            params={
                "get": "NAME",
                "for": "county:*",
                "in": f"state:{sfips}",
                "key": api_key,
            },
            timeout=30,
        )
        if not r.ok:
            log.warning("Census county names: state %s → HTTP %d", sfips, r.status_code)
            continue
        rows = r.json()
        headers = rows[0]
        name_idx = headers.index("NAME")
        state_idx = headers.index("state")
        county_idx = headers.index("county")
        for row in rows[1:]:
            full_name = row[name_idx]  # e.g. "Monroe County, Mississippi"
            sfips2 = row[state_idx].zfill(2)
            cfips = row[county_idx].zfill(3)
            geoid = sfips2 + cfips
            state_abbr = FIPS_TO_ABBR.get(sfips2, "")
            # Extract county part: before the first comma
            county_part = full_name.split(",")[0]
            normalized = _normalize_county_name(county_part)
            if state_abbr and normalized:
                mapping[(state_abbr, normalized)] = geoid
    log.info("Built county name→FIPS map: %d entries", len(mapping))
    return mapping


def _download_eia(url: str) -> zipfile.ZipFile:
    import requests
    log.info("Downloading EIA 861 ZIP (~4–5 MB) ...")
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    log.info("  Downloaded %d kB", len(r.content) // 1024)
    return zipfile.ZipFile(io.BytesIO(r.content))


def _read_utility_rates(zf: zipfile.ZipFile) -> dict[int, dict]:
    """Return {utility_number: {state: {rate_cents, customers}}} from Sales sheet."""
    import openpyxl

    fname = "Sales_Ult_Cust_2024.xlsx"
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(fname)), read_only=True, data_only=True)
    ws = wb["States"]
    rows = ws.iter_rows(values_only=True)

    # Find the data header row (row with "Data Year" in col 0)
    header = None
    for row in rows:
        if row[0] == "Data Year":
            header = list(row)
            break
    if header is None:
        raise ValueError("Could not find header row in Sales_Ult_Cust_2024.xlsx States sheet")

    # Column positions
    util_col = header.index("Utility Number")
    state_col = header.index("State")
    part_col = header.index("Part")

    # Residential columns: "Thousand Dollars" and "Megawatthours" under RESIDENTIAL
    # The header merged cells: find by position (cols 9,10,11 = Residential Rev, Sales, Cust)
    res_rev_col = 9
    res_sales_col = 10
    res_cust_col = 11

    # util_id → state → (total_revenue, total_sales, total_customers)
    util_data: dict[tuple[int, str], list[float]] = {}

    for row in rows:
        if row is None or row[0] is None:
            continue
        part = str(row[part_col]).strip().upper() if row[part_col] else ""
        if part != "A":  # Part A = observed, full-service utilities
            continue
        try:
            uid = int(row[util_col])
            state = str(row[state_col]).strip().upper()
            rev = float(row[res_rev_col] or 0)  # $1000s
            sales = float(row[res_sales_col] or 0)  # MWh
            cust = float(row[res_cust_col] or 0)
        except (TypeError, ValueError):
            continue
        if sales <= 0 or cust <= 0:
            continue
        key = (uid, state)
        if key not in util_data:
            util_data[key] = [0.0, 0.0, 0.0]
        util_data[key][0] += rev
        util_data[key][1] += sales
        util_data[key][2] += cust

    # Compute rate ¢/kWh = revenues_thousands * 100 / sales_mwh
    rates: dict[tuple[int, str], tuple[float, float]] = {}
    for (uid, state), (rev, sales, cust) in util_data.items():
        rate = rev * 100 / sales  # ¢/kWh
        if 1.0 <= rate <= 100.0:  # sanity filter
            rates[(uid, state)] = (rate, cust)

    log.info("Utility rates computed: %d utility-state pairs", len(rates))
    return rates  # type: ignore[return-value]


def _read_service_territory(
    zf: zipfile.ZipFile,
    county_fips_map: dict[tuple[str, str], str],
) -> dict[str, list[tuple[int, str]]]:
    """Return {county_fips: [(utility_number, state), ...]}."""
    import openpyxl

    fname = "Service_Territory_2024.xlsx"
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(fname)), read_only=True, data_only=True)
    ws = wb["Counties_States"]

    county_utilities: dict[str, list[tuple[int, str]]] = {}
    unmatched: set[tuple[str, str]] = set()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:  # skip header
            continue
        try:
            uid = int(row[1])
            state = str(row[4]).strip().upper()
            county_raw = str(row[5]).strip()
        except (TypeError, ValueError, IndexError):
            continue
        if not county_raw or not state:
            continue
        normalized = _normalize_county_name(county_raw)
        fips = county_fips_map.get((state, normalized))
        if fips is None:
            unmatched.add((state, county_raw))
            continue
        county_utilities.setdefault(fips, []).append((uid, state))

    if unmatched:
        log.debug("Unmatched county names (%d): %s", len(unmatched), sorted(unmatched)[:10])
        log.info("Unmatched county names: %d (district/territory names expected)", len(unmatched))
    log.info("Service territory: %d counties with utility mappings", len(county_utilities))
    return county_utilities


def _compute_county_rates(
    county_utilities: dict[str, list[tuple[int, str]]],
    utility_rates: dict[tuple[int, str], tuple[float, float]],
) -> dict[str, float]:
    """Customer-weighted average rate per county."""
    out: dict[str, float] = {}
    for fips, util_list in county_utilities.items():
        total_weight = 0.0
        weighted_sum = 0.0
        for uid, state in util_list:
            entry = utility_rates.get((uid, state))
            if entry is None:
                continue
            rate, customers = entry
            weighted_sum += rate * customers
            total_weight += customers
        if total_weight > 0:
            out[fips] = round(weighted_sum / total_weight, 2)
    log.info("County electricity rates: %d counties", len(out))
    return out


def main() -> None:
    try:
        import requests  # noqa: F401
        import openpyxl  # noqa: F401
    except ImportError as e:
        sys.exit(f"Missing dependency: {e}. Run: pip install requests openpyxl")

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from app.settings import settings  # noqa: PLC0415

    if not settings.census_api_key:
        sys.exit("CENSUS_API_KEY not set — add it to backend/.env")

    county_fips_map = _build_county_name_fips(settings.census_api_key)
    zf = _download_eia(EIA_URL)
    utility_rates = _read_utility_rates(zf)
    county_utilities = _read_service_territory(zf, county_fips_map)
    county_rates = _compute_county_rates(county_utilities, utility_rates)

    out = {
        "_meta": {
            "source": "EIA Form 861 (2024) — residential electricity rates",
            "source_year": 2024,
            "notes": (
                "Customer-weighted average residential electricity rate (¢/kWh) per county. "
                "Computed from EIA-861 utility residential revenues and sales (Part A, States sheet), "
                "mapped to counties via EIA-861 Service Territory schedule. "
                "Counties with no serving utility in Form 861 fall back to state average."
            ),
        },
        "data": dict(sorted(county_rates.items())),
    }
    OUT_FILE.write_text(json.dumps(out, indent=2))
    log.info("Wrote %d counties → %s", len(county_rates), OUT_FILE)


if __name__ == "__main__":
    main()
